import re
import os
import copy
import openpyxl
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import IconSetRule
from iap.forecasting.workbench.services import help_data_manager as data__service
from iap.forecasting.workbench.services.data_management import get_decomposition, transform_decomp_for_view


# TODO from iap.common.helper import dicts_left_join
def dicts_left_join_list(d1, d2):
    if isinstance(d2, list):
        for d2_dict in d2:
            for key in d1.keys():
                if key in d2_dict:
                    d1[key] = d2_dict[key]
                if 'lang-en-'+key in d2_dict:
                    d1[key] = d2_dict['lang-en-'+key]
    else:
        for key in d1.keys():
            if key in d2:
                d1[key] = d2[key]


def get_options_data(config, scenarios):
    """
    Returns options for report

    :param config:
    :type config: work_bench.data_config
    :param scenarios:
    :type scenarios: list
    :return:
    :rtype: dict
    """
    # Initialize structure for output.
    options = dict(
        timescales=None,
        primary_only=None,
        scenarios=None,
        period=None
    )

    options['timescales'] = dict()
    options['primary_only'] = 1
    options['scenarios'] = scenarios  # [{'scenario_name_1': 1}]  # config.get_property('report_scenarios')
    options['period'] = ['2012', '2018']  # config.get_property('dash_top_ts_period')
    # Fill timescales
    periods = config.get_property('dash_timescales')  # ['annual']

    # TODO getting settings from warehouse (?WB)
    sections = ['forecast', 'coefficients', 'volume_decomposition', 'value_decomposition']
    # Fill first level
    for period in periods:
        options['timescales'][period] = dict()
    # Fill second level
    for period in periods:
        for section in sections:
            if section in ['decomposition']:
                options['timescales'][period][section] = ['annual']  # TODO
            else:
                options['timescales'][period][section] = 1
    return options


def collect_report_data(options, permission_tree, container, config, entities_ids, lang):
    """
    Returns data for report

    :param options: report options from user
    :type options: dict
    :param permission_tree:
    :type permission_tree: dict
    :param container: from user's workbench
    :type container: work_bench.default_container
    :param config: from user's workbench
    :type config: work_bench.data_config
    :param lang: language
    :type lang: str
    :return:
    :rtype: dict
    """
    # Output list
    report_structure_list = list()
    # Add default scenario
    scenario_list = [{'name': 'Default', 'status': 'Final'}, {'name': 'Scenario_1', 'status': 'Final'},
                     {'name': 'Scenario_2', 'status': 'Final'}]  # + options['scenarios']
    default_header = ['Country', 'Category', 'Product']

    for scenario in scenario_list:
        if scenario['status'] != 'Final':
            continue
        report_structure_data = dict(
            data=list(),
            config=dict()
        )
        report_structure_data['config']['scenario_info'] = scenario
        report_structure_data['config']['timescale'] = list(options['timescales'].keys())[0]

        for ids in range(1, len(container._nodes_dict)):
            entities_ids = [ids]
            # Structure for report.
            report_data = dict(
                data=dict(
                    section=None,
                    timescales=None,
                    variables=None,
                    timelabels=None,
                    variable_values=None,
                    factor_drivers=None,
                    coefficients=None,
                    decomp=None,
                    change_over_period=None,
                    decomp_types=None,
                    decomp_type_factors=None,
                    insights=None
                ),
                config=None
            )

            ent = container.get_entity_by_id(ids)

            report_data['config'] = dict(
                scenario_name=scenario['name'],
                scenario_info=scenario
            )

            for period in options['timescales']:
                period_options = options['timescales'][period]

                main_timescales = list(options['timescales'].keys())  # config.get_property('dash_timescales')
                top_ts_period = options['period']
                dec_timescales = config.get_property('dash_decomposition_timescales')  # period_options['decomposition']
                top_ts = str(main_timescales[0])
                bottom_ts = str(main_timescales[-1])

                mid = container.timeline.get_period_by_alias(top_ts, 'history')[0][1]
                report_data['config']['decomp_timescales'] = [str(x) for x in dec_timescales]
                report_data['config']['main_period'] = dict(timescale=top_ts, start=top_ts_period[0], mid=mid,
                                                            end=top_ts_period[1]),
                report_data['config']['decomp_period'] = dict(timescale=top_ts, start=mid, end=top_ts_period[1])

                # Get timelabels tree and time borders for every timescale.
                ts_tree, ts_borders = container.timeline.get_timeline_tree(top_ts, bottom_ts, top_ts_period)
                report_data['data']['timelabels'] = list(ts_tree)

                # Get timescales view settings.
                timescales_info = config.get_objects_properties('timescale', main_timescales, lang)
                timescales_view_info = []

                for ts_info in timescales_info:
                    ts_view_props = dict(
                        id=None,
                        full_name=None,
                        short_name=None,
                        lag=None
                    )
                    dicts_left_join_list(ts_view_props, ts_info)
                    ts_view_props['lag'] = \
                        container.timeline.get_growth_lag(ts_info[0]['id'])
                    timescales_view_info.append(ts_view_props)
                report_data['data']['timescales'] = list(timescales_view_info)

                # Get time labels for every timescale.
                # Get growth rates periods for every timescale.
                time_labels = dict()
                gr_periods = dict()
                for ts_name, ts_period in ts_borders.items():
                    time_labels[ts_name] = container.timeline.get_names(ts_name, ts_period)
                    gr_periods[ts_name] = container.timeline.get_growth_periods(ts_name, ts_period)
                    gr_periods[ts_name].extend(
                        container.timeline.get_carg_periods(ts_name, ts_period)
                    )

                # Get variables to view. Values, growth rates, CAGRS.
                vars_view_props = []
                time_series_data = dict([(x, dict()) for x in ts_borders.keys()])
                periods_data = dict([(x, dict()) for x in ts_borders.keys()])

                try:
                    items_view_props = config.get_vars_for_view(meta=ent.meta, path=ent.path)
                except Exception:
                    continue
                    # TODO change on no var for view

                for item in items_view_props:
                    # Get entity to get variables from.
                    curr_ent = container.get_entity_by_filter(ent, item['filter'])
                    # Collect variables data.
                    absent_vars_ids = []
                    var_items = item['variables']
                    for var_info in var_items:
                        var_id = var_info['id']
                        var = curr_ent.get_variable(var_id)
                        if var is None:
                            absent_vars_ids.append(var_info['id'])
                            continue

                        data__service.get_time_series_values(permission_tree=permission_tree, ent=ent,
                                                             ts_borders=ts_borders,
                                                             var=var, periods_data=periods_data,
                                                             time_series_data=time_series_data,
                                                             var_info=var_info, gr_periods=gr_periods)
                    # Get variables properties.
                    vars_ids = [var_info['id']
                                for var_info in item['variables']
                                if var_info['id'] not in absent_vars_ids]
                    vars_types = [var_info['type']
                                  for var_info in item['variables']
                                  if var_info['id'] not in absent_vars_ids]

                    vars_props = config.get_objects_properties('variable', vars_ids, lang)

                    for index, v_props in enumerate(vars_props):
                        view_props = dict(
                            id=None,
                            full_name=None,
                            short_name=None,
                            type=None,
                            metric=None,
                            format=None,
                            multiplier=None,
                            hint=''
                        )
                        dicts_left_join_list(view_props, v_props)
                        view_props['type'] = vars_types[index]
                        vars_view_props.append(view_props)
                report_data['data']['variable_values'] = dict(time_series_data)

                # Get decomposition types properties.
                decs_types_view_props = data__service.get_decs_types_view_props(config, lang)
                report_data['data']['decomp_types'] = decs_types_view_props

                # Get decomposition.
                dec_periods = {key: value for key, value in gr_periods.items()
                               if key in dec_timescales}

                decomp_data = get_decomposition(container, config, entities_ids, dec_periods)
                decomp_data_for_view = transform_decomp_for_view(decomp_data)

                # Update periods data with decomposition
                for row in decomp_data:
                    if row['var_id'] not in periods_data[row['ts_name']]:
                        periods_data[row['ts_name']][row['var_id']] = []
                    periods_data[row['ts_name']][row['var_id']] = (dict(abs=row['abs'], rate=row['rate'],
                                                                        start=row['start'], end=row['end']))

                report_data['data']['decomp'] = decomp_data_for_view
                report_data['data']['change_over_period'] = periods_data

                # Fill factors for dec type.
                dec_factors = data__service.get_factors_for_dec_type(decomp_data_for_view=decomp_data_for_view)
                report_data['data']['decomp_type_factors'] = dec_factors

                # Extend variables properties.
                for item in config.get_decomp_vars_for_view(meta=ent.meta, path=ent.path):
                    vars_ids = [var_info['id'] for var_info in item['variables']]
                    vars_props = \
                        config.get_objects_properties('variable', vars_ids, lang)
                    for index, v_props in enumerate(vars_props):
                        view_props = dict(id=None, full_name=None, short_name=None,
                                          type=None, metric=None, format=None, hint='')
                        dicts_left_join_list(view_props, v_props)
                        view_props['type'] = 'impact'
                        vars_view_props.append(view_props)
                report_data['data']['variables'] = vars_view_props

                # Relations between factors and drivers
                factor_drivers = config.get_factor_drivers_relations(meta=ent.meta, path=ent.path)
                for factor, fd_rel in factor_drivers.items():
                    factor_drivers[factor] = \
                        [dict(factor=x[0], driver=x[1]) for x in fd_rel]
                report_data['data']['factor_drivers'] = list(factor_drivers)

                # Get Insights.
                report_data['data']['insights'] = list(dict(text=x) for x in ent.insights)

                # Get scalars(coefficients)
                report_data['data']['coefficients'] = dict(ent._data._scalars)

                report_structure = dict()
                for section in period_options:
                    section_dict = dict(
                        path=None,
                        rows=[],
                        format=[],
                        header_format=None
                    )
                    section_dict['path'] = ent.path

                    # Section coefficients
                    if section == 'coefficients':
                        values_row = section_dict['path']
                        header = default_header[:len(values_row)]

                        for coefficient in report_data['data']['coefficients']:
                            if coefficient[1] == report_data['config']['main_period'][0]['timescale']:
                                header.append(coefficient[0].replace('_', ' ').title())
                                if isinstance(report_data['data']['coefficients'][coefficient], list):
                                    values_row.append(report_data['data']['coefficients'][coefficient][0])
                                else:
                                    values_row.append(report_data['data']['coefficients'][coefficient])
                        section_dict['rows'].append(header)
                        section_dict['rows'].append(values_row)

                    # Section forecast and decomposition
                    else:
                        data_to_transform = None
                        if section == 'forecast':
                            data_to_transform = report_data['data']['variable_values']
                        if section == 'value_decomposition':
                            data_to_transform = report_data['data']['decomp']['annual']['value']
                        if section == 'volume_decomposition':
                            data_to_transform = report_data['data']['decomp']['annual']['volume']

                        values_row_start = section_dict['path']
                        header = default_header[:len(values_row_start)]
                        if data_to_transform is not None:
                            values_list = data_to_table(section,
                                                        report_data['data']['timelabels'],
                                                        data_to_transform,
                                                        report_data['data']['variables'])
                        for n_row in range(0, len(values_list['rows'])):
                            if n_row == 0:
                                values_row = header + values_list['rows'][n_row]
                            else:
                                values_row = values_row_start + values_list['rows'][n_row]
                            section_dict['rows'].append(values_row)
                            if section == 'forecast':
                                section_dict['format'].append(values_list['format'][n_row])
                        # Header format
                        if section == 'forecast':
                            if len(values_list['rows'])>0:
                                header = default_header[:len(values_row_start)] + values_list['rows'][0]
                                start = ent._data.time_manager._period_alias['forecast']['annual'][0]
                                section_dict['header_format'] = [len(values_row_start)+2, header.index(start)]
                        else:
                            section_dict['header_format'] = [len(values_row_start)+1]

                    report_structure[section] = dict(section_dict)

                report_structure_data['data'].append(copy.deepcopy(report_structure))
        report_structure_list.append(copy.deepcopy(report_structure_data))
    return report_structure_list


def data_to_table(section, time_labels, variable_values, variables):
    """ Returns structure for excel report

    """
    result_table = dict(
        rows=list(),
        format=list()
    )
    try:
        if section == 'forecast':
            mask = ['Fact', 'Metric']
            for time_label in time_labels:
                mask.append(time_label['full_name'])

            result_table = {'rows': [mask], 'format': [None]}
            for timescale in variable_values:
                var_list = variable_values[timescale]
                for var in var_list:
                    table = {}
                    for num in range(0, len(var_list[var]['values'])):
                        table[var_list[var]['stamps'][num]] = var_list[var]['values'][num]
                    var_info = next((item for item in variables if item['id'] == var), None)
                    line = [var_info['full_name'], var_info['metric']]
                    for num in range(2, len(mask)):
                        line.append(0)
                        if mask[num] in table and var_info['multiplier'] != '':
                            line[num] = table[mask[num]]*float(var_info['multiplier'])
                    result_table['rows'].append(line)
                    result_table['format'].append(var_info['format'].replace(",", ".").replace("'", ""))
        else:
            header = ['Due to factor']
            for i in variable_values:
                header.append(i['start'][2:] + "/" + i['end'][2:])
            result_table['rows'].append(header)

            for i in variable_values[0]['factors']:
                var_info = next((item for item in variables if item['id'] == i['var_id']), None)
                line = [var_info['full_name']]
                for var_date in variable_values:
                    var_info = next((item for item in var_date['factors'] if item['var_id'] == i['var_id']), None)
                    line.append(var_info['rate'])
                result_table['rows'].append(line)
    except Exception:
        result_table['rows'] = list()
        result_table['format'] = list()
    return result_table


""" ===================================================================================================================
EXCEL part
"""


def styled_row(sheet, data, number_format=None, bold=False, color=None, color_txt=None):
    """ Adds styles to row data before writing

    """
    for cell in data:
        cell = Cell(sheet, column="A", row=1, value=cell)
        if number_format is not None:
            cell.number_format = number_format
        if bold:
            cell.font = Font(bold=True)
        if color is not None:
            cell.fill = PatternFill(start_color=color, fill_type='solid')
        if color_txt is not None:
            cell.font = Font(color=color_txt)
        yield cell


def create_report(file_name, report_data_list):
    """ Generates xlsx report to file_name

    """
    if file_name is None:
        file_name = os.path.expanduser("~/Desktop/sample.xlsx")

    color_scenario = '00b050'
    color_timescale = 'ffff00'
    color_section = 'c5d9f1'
    color_coefficient = '808080'
    color_coefficient_txt = 'ffffff'
    color_cagr = 'fde9d9'
    color_history = 'b8cce4'
    color_forecast = 'c0504d'
    rule = IconSetRule(icon_style='5Arrows', type='num', values=[0, -0.005, -0.0001, 0.0001, 0.005], percent=False)

    work_book = openpyxl.Workbook()
    sheet = work_book.active
    sht = 0

    for scenario in report_data_list:
        row = 1
        clm = 1
        # Sheet name
        scenario_name = scenario['config']['scenario_info']['name'][:31]  # Excel sheet title limit 31 char
        if sht != 0:
            work_book.create_sheet(scenario_name)
            sheet = work_book.get_sheet_by_name(scenario_name)
        else:
            sheet.title = scenario_name
        # Scenario
        sheet.cell(row=row, column=clm).value = scenario['config']['scenario_info']['name'].capitalize()
        sheet.cell(row=row, column=clm).fill = PatternFill(start_color=color_scenario, fill_type='solid')
        sheet.cell(row=row, column=clm).font = Font(bold=True)
        row += 1
        # Timescale
        sheet.cell(row=row, column=clm).value = scenario['config']['timescale'].capitalize()
        sheet.cell(row=row, column=clm).fill = PatternFill(start_color=color_timescale, fill_type='solid')
        sheet.cell(row=row, column=clm).font = Font(bold=True)
        row += 1
        for entity in scenario['data']:
            # Order of sections output
            for section in ['forecast', 'coefficients', 'value_decomposition', 'volume_decomposition']:
                if section in entity:
                    section_data = entity[section]
                    if len(section_data['rows']) == 0:
                        continue
                    # Name of section
                    sheet.cell(row=row, column=clm).value = section.replace('_', ' ').title()
                    sheet.cell(row=row, column=clm).fill = PatternFill(start_color=color_section, fill_type='solid')
                    sheet.cell(row=row, column=clm).font = Font(bold=True)
                    row += 1
                    # Table header
                    if section == 'coefficients':
                        sheet.append(styled_row(sheet, section_data['rows'][0], color=color_coefficient,
                                                color_txt=color_coefficient_txt))
                    else:
                        sheet.append(styled_row(sheet, section_data['rows'][0], bold=True))
                        if section == 'forecast':
                            for i in range(section_data['header_format'][0], section_data['header_format'][1]):
                                sheet.cell(row=row, column=i+1).fill = PatternFill(start_color=color_history,
                                                                                   fill_type='solid')
                            for i in range(section_data['header_format'][1], len(section_data['rows'][0])):
                                sheet.cell(row=row, column=i+1).fill = PatternFill(start_color=color_forecast,
                                                                                   fill_type='solid')
                        else:
                            for i in range(section_data['header_format'][0], len(section_data['rows'][0])):
                                sheet.cell(row=row, column=i+1).fill = PatternFill(start_color=color_cagr,
                                                                                   fill_type='solid')
                    row += 1
                    # Table data
                    for i in range(1, len(section_data['rows'])):

                        complex_flag = 0
                        for nun in section_data['rows'][i]:
                            if isinstance(nun, complex):
                                complex_flag = 1
                        if complex_flag:
                            continue

                        if section == 'forecast':
                            sheet.append(styled_row(sheet, section_data['rows'][i],
                                                    number_format=section_data['format'][i]))
                        elif section == 'coefficients':
                            sheet.append(section_data['rows'][i])
                        else:
                            sheet.conditional_formatting.add(str(row) + ':' + str(row), rule)
                            sheet.append(styled_row(sheet, section_data['rows'][i], number_format='0.#%'))
                        row += 1
                row += 1  # before next section
            row += 3  # before next scenario

        # Adjust width
        column_range = re.sub(r'\d+', '', sheet.dimensions).split(':')
        column_list = ''.join(chr(c) for c in range(ord(column_range[0]), ord(column_range[1]) + 1))
        for col in column_list:
            sheet.column_dimensions[col].width = 20

        sht += 1  # before next sheet

    # Save the file
    work_book.save(file_name)
